
import json
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
import threading
import time
import threading

#线程数
num_threads = 4

# 定义线程处理函数
def process_urls(urls,res):
    for url in urls:
        # 处理每个url的逻辑
        for data in url:
            getdata(data.strip(),res)


def getfiledata():
    # with open('file.txt', 'r') as file:
    #     line = file.readline()
    #     while line:
    #         # print(line)
    #         getdata(line.strip())
    #         line = file.readline()
    # 分块读取文件中的所有url
    with open('file.txt', 'r') as f:
        chunk_size = 50  # 每次读取的最大行数
        chunk = []
        chunks = []
        for i, line in enumerate(f):
            i+=1
            chunk.append(line)
            if i % (chunk_size) == 0:
                chunks.append(chunk)
                chunk=[]
        if(chunk!=[]):
            chunks.append(chunk)
            chunk=[]
    # 将每个子列表分配给一个线程处理
    
    
    threads = []
    if(i<num_threads*chunk_size):
        num_threads=2
    if(i<100):
        num_threads=1
    for i in range(num_threads):
        start=round(i * float(len(chunks)/num_threads))
        end=round((i+1) * float(len(chunks)/num_threads))
        if(end>len(chunks)+1):
            end=len(chunks)+1
        t = threading.Thread(target=process_urls, args=(chunks[start:end],"线程"+str(i)))
        t.start()
        threads.append(t)
    # print(threads)
    # 等待所有线程结束
    for t in threads:
        t.join()


# 前后字符串
start_str = "<div class=\"table-responsive\">"
end_str = "<form method=\"post\" class=\"hidden\" action=\"/\">"
#line在上边去除过\n,\n是在添加文件是添加的
def getdata(line,res):
    html_response=requests.get(line)
    html_resp=html_response.text
    # 查找前后字符串在原始字符串的位置索引
    start_index = html_resp.find(start_str)
    end_index = html_resp.find(end_str)
    end_index = html_resp.find(end_str)
    # 判断前后字符串是否都在原始字符串中
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = html_resp[start_index+len(start_str):end_index]
        # 输出结果
        # print(result_str)
    else:
        # print("前后")
        print("前后字符串不存在或有误！")
    # print("-------------------")
    # print(result_str is not None)
    if result_str is not None:
        setExcel(result_str,res)

# 创建一个线程锁
lock = threading.Lock()
now_row=2
def setExcel(result_str,res):
    global now_row
    soup = BeautifulSoup(result_str, 'html.parser')
    sci_name = soup.select_one('.details_species_text').text.strip()
    chinesename=getchinesename(result_str)
    addr=getaddr(result_str)
    addr_english=getaddr_english(result_str)
    datasource=getdatasource(result_str)
    types=gettypes(result_str)
    #向excel中写入对应的value
    # print("学名："+sci_name)
    # print("分布地："+addr_english)
    # print("分布地中文:"+addr)
    # print("中文名："+chinesename)
    # print("数据来源："+datasource)
    # print(types)
    # 获取线程锁
    lock.acquire()
    try:
        sheet.cell(row=now_row, column=1).value = addr_english
        sheet.cell(row=now_row, column=2).value = addr
        sheet.cell(row=now_row, column=3).value = sci_name
        sheet.cell(row=now_row, column=4).value = chinesename
        sheet.cell(row=now_row, column=5).value = datasource
        sheet.cell(row=now_row, column=6).value = types[0]
        sheet.cell(row=now_row, column=7).value = types[1]
        sheet.cell(row=now_row, column=8).value = types[2]
        sheet.cell(row=now_row, column=9).value = types[3]
        sheet.cell(row=now_row, column=10).value = types[4]
        sheet.cell(row=now_row, column=11).value = types[5]
        sheet.cell(row=now_row, column=13).value = res
        wb.save('物种爬取测试.xlsx')
    finally:
        # 释放线程锁
        lock.release()
    now_row+=1
    # print('数据写入成功！')
    print(now_row)
    # if(now_row>=100):
    #     exit()

    

def getchinesename(result_str):
    # 查找前后字符串在原始字符串的位置索引
    start_index = result_str.find("中文名")
    end_index = result_str.find("异名")
    # 判断前后字符串是否都在原始字符串中
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = result_str[start_index+len(start_str):end_index]
        # print(result_str)
        # 定义正则表达式
        pattern = re.compile(r'-->')
        # 输出结果
        chinese_name = re.sub(r"[^\u4e00-\u9fa5]", '', result_str)  # 删除HTML标签和注释
        # 进行替换操作
        chinese_name = re.sub(pattern, '', chinese_name)
        chinese_name = re.sub(r'\s+', ' ', chinese_name)  # 删除多余空格
        # print(chinese_name)
        return chinese_name
    else:
        # print("前后字符串不存在或有误！")
        return None   

def getaddr(result_str):
    # 查找前后字符串在原始字符串的位置索引
    start_index = result_str.find("分布地(中文)")
    end_index = result_str.find("显示分布图")
    # 判断前后字符串是否都在原始字符串中
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = result_str[start_index+len(start_str):end_index]
        soup = BeautifulSoup(result_str, 'html.parser')
        td_element = soup.select_one('td.details_text sapn')
        addrname = td_element.text.strip()
        return addrname
    else:
        # print("前后字符串不存在或有误！")
        return None

def getaddr_english(result_str):
    # 查找前后字符串在原始字符串的位置索引
    start_index = result_str.find("分布地")
    end_index = result_str.find("分布地(中文)")
    # 判断前后字符串是否都在原始字符串中
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = result_str[start_index+len(start_str):end_index]
        soup = BeautifulSoup(result_str, 'html.parser')
        td_element = soup.select_one('td.details_text sapn')
        addrname_english = td_element.text.strip()
        # print(addrname_english)
        return addrname_english
    else:
        # print("前后字符串不存在或有误！") 
        return None  

def getdatasource(result_str):
    # 查找前后字符串在原始字符串的位置索引
    start_index = result_str.find("源数据库")
    end_index = result_str.find("审核专家")
    if(end_index==-1):
        end_index = result_str.find("收录日期")
    # 判断前后字符串是否都在原始字符串中
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = result_str[start_index+len(start_str):end_index]
        soup = BeautifulSoup(result_str, 'html.parser')
        td_element = soup.select_one('a')
        datasource = td_element.text.strip()
        # print(datasource)
        return datasource
    else:
        # print("前后字符串不存在或有误！") 
        return None 

#获取所有的界门等
def gettypes(result_str):
    # 查找前后字符串在原始字符串的位置索引
    start_index = result_str.find("分类系统")
    end_index = result_str.find("分布地")
    if(end_index==-1):
        end_index = result_str.find("附加信息")
    # 判断前后字符串是否都在原始字符串中
    list=['','','','','','']
    if start_index != -1 and end_index != -1:
        # 切割字符串
        result_str = result_str[start_index+len(start_str):end_index]
        soup = BeautifulSoup(result_str, 'html.parser')
        array = []
        for li in soup.find_all('li'):
            data = li.find_all('sapn')[1].text
            array.append(data)
        # print(array)
        return array
    else:
        # print("前后字符串不存在或有误！")
        return list

    



if __name__ == '__main__':
    excel_path="中国生物物种名录.xlsx"
    wb = Workbook()
    sheet = wb.active
    #向excel中写入表头
    sheet['a1'] = '分布地'
    sheet['b1'] = '分布地中文'
    sheet['c1'] = '接收的学名'
    sheet['d1'] = '中文名'
    sheet['e1'] = '数据来源'
    sheet['f1'] = '所属界名'
    sheet['g1'] = '所属门名'
    sheet['h1'] = '所属纲名'
    sheet['i1'] = '所属目名'
    sheet['j1'] = '所属科名'
    sheet['k1'] = '所属属名'
    sheet['m1'] = '线程号：'
    #获取文件中的所有url。
    getfiledata()
    # 保存 Excel 文件
    wb.save('中国生物物种名录.xlsx')
    